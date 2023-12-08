from flask import Blueprint, Response
from models import Usuario, Cafe, Venta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

appcsv = Blueprint("appcsv", __name__, template_folder="templates")

def generateExcel(usuarios, cafes, ventas):
    # Crear un libro de trabajo
    wb = Workbook()
    # Seleccionar la primera hoja
    ws = wb.active

    # Establecer estilos para el encabezado
    header_style = PatternFill(start_color="00FFCCFF", end_color="00FFCCFF", fill_type="solid")
    font_style = Font(bold=True)

    # Encabezados para usuarios
    ws.append(["ID", "EMAIL", "REGISTRADO", "ADMIN", "PASSWORD"])
    for cell in ws[1]:
        cell.fill = header_style
        cell.font = font_style

    # Datos para usuarios
    for usuario in usuarios:
        ws.append([usuario.id, usuario.email, usuario.registrado, usuario.admin, usuario.password])

    # Encabezados para cafés
    ws.append(["ID", "NOMBRE", "PRECIO", "INGREDIENTES"])
    for cell in ws[ws.max_row]:
        cell.fill = header_style
        cell.font = font_style

    # Datos para cafés
    for cafe in cafes:
        ws.append([cafe.id, cafe.nombre, cafe.precio, cafe.ingredientes])

    # Encabezados para ventas
    ws.append(["ID","CLIENTE","CAFES","CANTIDAD","TOTAL"])
    for cell in ws[ws.max_row]:
        cell.fill = header_style
        cell.font = font_style

    # Datos para ventas
    for venta in ventas:
        ws.append([venta.id, venta.cliente, venta.cafes, venta.cantidad, venta.total])

    # Guardar el libro de trabajo en un búfer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    headers = {
        "Content-Disposition": "attachment;filename=datos.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

    return Response(excel_buffer.read(), headers=headers)

# Asocia la ruta /excel con la función generateExcel
@appcsv.route('/csv')
def excel():
    usuarios = Usuario.query.all()
    cafes = Cafe.query.all()
    ventas = Venta.query.all()
    return generateExcel(usuarios, cafes, ventas)
